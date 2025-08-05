from django.http import HttpResponse, HttpResponseServerError
from datetime import datetime
from .models import CustomUser

def download_user_pdf_excel_combined(modeladmin, request, queryset):
    """Export selected users - attempts PDF first, then Excel fallback"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
    except ImportError:
        # Fallback to Excel if reportlab is not installed
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Return error if neither library is available
            return HttpResponseServerError("Either ReportLab (for PDF) or openpyxl (for Excel) is required. Please install one of these packages.")
        
        # Excel generation if openpyxl is available
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="selected_users.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Selected Users"
        
        # Title
        ws['A1'] = "Nestorc Selected Users Report"
        title_font = Font(size=16, bold=True, color="366092")
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:E1')
        
        # Metadata
        ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Users: {queryset.count()}"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add headers
        headers = ['Full Name', 'Email', 'Phone Number', 'Is Active', 'Is Staff']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, user in enumerate(queryset, 7):
            ws.cell(row=row, column=1, value=user.full_name or 'N/A')
            ws.cell(row=row, column=2, value=user.email)
            ws.cell(row=row, column=3, value=user.phone_number or 'N/A')
            ws.cell(row=row, column=4, value="Yes" if user.is_active else "No")
            ws.cell(row=row, column=5, value="Yes" if user.is_staff else "No")
        
        # Auto-adjust column widths (avoid merged cells issue)
        column_letters = ['A', 'B', 'C', 'D', 'E']
        for i, col_letter in enumerate(column_letters, 1):
            max_length = 0
            
            # Check header length
            header_cell = ws.cell(row=6, column=i)
            if header_cell.value:
                max_length = len(str(header_cell.value))
            
            # Check all data cells in this column
            for row_num in range(7, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=i)
                if cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    # Enhanced PDF generation if reportlab is available
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="selected_users.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    # Header
    p.setFont("Helvetica-Bold", 20)
    p.drawString(100, height - 100, "Selected Users Report")
    
    p.setFont("Helvetica", 12)
    p.drawString(100, height - 130, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    p.drawString(100, height - 150, f"Total Users: {queryset.count()}")
    
    # Draw a line
    p.line(100, height - 170, width - 100, height - 170)
    
    # Table headers
    y = height - 200
    p.setFont("Helvetica-Bold", 10)
    p.drawString(100, y, "Full Name")
    p.drawString(250, y, "Email")
    p.drawString(400, y, "Phone")
    p.drawString(500, y, "Active")
    
    # Table content
    p.setFont("Helvetica", 9)
    y -= 20
    
    for user in queryset:
        if y < 50:  # Start new page if near bottom
            p.showPage()
            y = height - 100
            # Repeat headers on new page
            p.setFont("Helvetica-Bold", 10)
            p.drawString(100, y, "Full Name")
            p.drawString(250, y, "Email")
            p.drawString(400, y, "Phone")
            p.drawString(500, y, "Active")
            p.setFont("Helvetica", 9)
            y -= 20
        
        # Truncate long text to fit
        full_name = (user.full_name or 'N/A')[:20]
        email = user.email[:25] if len(user.email) > 25 else user.email
        phone = (user.phone_number or 'N/A')[:15]
        active_status = "Yes" if user.is_active else "No"
        
        p.drawString(100, y, full_name)
        p.drawString(250, y, email)
        p.drawString(400, y, phone)
        p.drawString(500, y, active_status)
        y -= 15

    p.showPage()
    p.save()
    return response

download_user_pdf_excel_combined.short_description = "Download PDF/Excel of selected users"
