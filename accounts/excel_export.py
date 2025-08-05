from django.http import HttpResponse
from datetime import datetime
from .models import CustomUser

def download_user_excel(queryset):
    """Export selected users to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="selected_users.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Full Name', 'Email', 'Phone Number', 'Is Active', 'Is Staff'])
        
        for user in queryset:
            writer.writerow([
                user.full_name or '',
                user.email,
                user.phone_number or '',
                user.is_active,
                user.is_staff
            ])
        return response
    
    # Excel generation if openpyxl is available
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="selected_users.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Selected Users"
    
    # Title
    ws['A1'] = "Nestorc Selected Users Report"
    title_font = Font(size=16, bold=True, color="366092")
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:E1')
    
    # Metadata
    ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Total Users: {queryset.count()}"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add headers
    headers = ['Full Name', 'Email', 'Phone Number', 'Is Active', 'Is Staff']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, user in enumerate(queryset, 7):
        ws.cell(row=row, column=1, value=user.full_name or 'N/A')
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=user.phone_number or 'N/A')
        ws.cell(row=row, column=4, value="Yes" if user.is_active else "No")
        ws.cell(row=row, column=5, value="Yes" if user.is_staff else "No")
    
    # Auto-adjust column widths (avoid merged cells issue)
    column_letters = ['A', 'B', 'C', 'D', 'E']
    for i, col_letter in enumerate(column_letters, 1):
        max_length = 0
        
        # Check header length
        header_cell = ws.cell(row=6, column=i)
        if header_cell.value:
            max_length = len(str(header_cell.value))
        
        # Check all data cells in this column
        for row_num in range(7, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=i)
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(response)
    return response

download_user_excel.short_description = "Download Excel of selected users"


def download_all_users_excel(request):
    """Download all users as Excel"""
    queryset = CustomUser.objects.all()
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="all_users.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Full Name', 'Email', 'Phone Number', 'Is Active', 'Is Staff'])
        
        for user in queryset:
            writer.writerow([
                user.full_name or '',
                user.email,
                user.phone_number or '',
                user.is_active,
                user.is_staff
            ])
        return response
    
    # Excel generation
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="nestorc_all_users.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "All Users Report"
    
    # Title
    ws['A1'] = "Nestorc Complete Users Report"
    title_font = Font(size=18, bold=True, color="366092")
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:E1')
    
    # Metadata
    ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Total Users: {queryset.count()}"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add headers
    headers = ['Full Name', 'Email', 'Phone Number', 'Is Active', 'Is Staff']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, user in enumerate(queryset, 7):
        ws.cell(row=row, column=1, value=user.full_name or 'N/A')
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=user.phone_number or 'N/A')
        ws.cell(row=row, column=4, value="Yes" if user.is_active else "No")
        ws.cell(row=row, column=5, value="Yes" if user.is_staff else "No")
    
    # Auto-adjust column widths (avoid merged cells issue)
    column_letters = ['A', 'B', 'C', 'D', 'E']
    for i, col_letter in enumerate(column_letters, 1):
        max_length = 0
        
        # Check header length
        header_cell = ws.cell(row=6, column=i)
        if header_cell.value:
            max_length = len(str(header_cell.value))
        
        # Check all data cells in this column
        for row_num in range(7, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=i)
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(response)
    return response

download_all_users_excel.short_description = "Download all users as Excel"
